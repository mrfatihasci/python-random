
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
a="/media/monster/New Volume/HOME/2WRITE/2Vocabulary excel/vocabulary4.xlsx"
wb=load_workbook("/media/monster/New Volume/HOME/2WRITE/2Vocabulary excel/vocabulary4.xlsx")
ws=wb.active
max_rows =ws.max_row +1
#print(max_rows)
import pyttsx3
speaker = pyttsx3.init()    # you might wanna write "espeak" inside the paranthesis
speaker.setProperty('rate', 120)
speaker.setProperty('volume',0.5)     # this floating value of volume varies in the range of 0.0 to 1.0
#voices = speaker.getProperty('voices')
#speaker.setProperty('voice', voices[1].id)
from googletrans import Translator, constants
from pprint import pprint
translator = Translator()                      # init the Google API translator
speaker.say("welcome   ")
speaker.runAndWait()
c=1
d=0
for i in range(2,max_rows):
	if(ws[str.format("E"+str(i))].value != c):
		d+=1
		cell1=ws[str.format("A"+str(i))].value
		cell1=cell1.upper()
		cell2=ws[str.format("B"+str(i))].value
		cell2=cell2.lower()
		cell2=str(cell2)
		cell2=cell2.strip()
		cell3 = translator.translate(cell1, src='en', dest='tr')
		#cell3=cell3.lower()
		cell4=ws[str.format("D"+str(i))].value
		print(" ")
		print(" ")
		print(" ")
		print(" ")
		print(" ")
		print(" ")
		print(" ")
		print(" ")
		print("                                                               ",d,"-) ",cell4,"             ")
		print(" ")
		print(" ")
		print("                                                                     ",cell1,":",end = "")
		guess=input()
		if guess==cell2:
			ws[str.format("E"+str(i))]=1
			wb.save(a)
			speaker.say(cell1)
			speaker.runAndWait()
			print(" ")
			print(" ")
			print("                                                                      Google : ",cell1,"=",cell3.text.lower(),"   ")
			print(" ")
			print(" ")
			print("                                                                     ","\u2713" )
			ws[str.format("C"+str(i))]=cell3.text.lower()
			wb.save(a)
			print(" ")
			print(" ")
			print("                                                                      Click enter to continue")
			clear=input()
			print("\033c", end="")
		else:
			print(" ")
			print(" ")
			print("                                                                     ",cell1,":",end = "")
			guess=input()
			if guess==cell2:
				ws[str.format("E"+str(i))]=1
				wb.save(a)
				speaker.say(cell1)
				speaker.runAndWait()
				print(" ")
				print(" ")
				print("                                                                      Google : ",cell1,"=",cell3.text.lower(),"   ")
				print(" ")
				print(" ")
				print("                                                                     ","\u2713" )
				ws[str.format("C"+str(i))]=cell3.text.lower()
				wb.save(a)
				print(" ")
				print(" ")
				print("                                                                      Click enter to continue")
				clear=input()
				print("\033c", end="")
			else:
				print(" ")
				print(" ")
				print("                                                                     ",cell1,":",end = "")
				guess=input()
				if guess==cell2:
					ws[str.format("E"+str(i))]=1
					wb.save(a)
					speaker.say(cell1)
					speaker.runAndWait()
					print(" ")
					print(" ")
					print("                                                                      Google : ",cell1,"=",cell3.text.lower(),"   ")
					print(" ")
					print(" ")
					print("                                                                     ","\u2713" )
					ws[str.format("C"+str(i))]=cell3.text.lower()
					wb.save(a)
					print(" ")
					print(" ")
					
					print("                                                                      Click enter to continue")
					clear=input()
					print("\033c", end="")
				else:
					print(" ")
					print(" ")
					print("                                                                     ",cell1,":",cell2,"       Google : ",cell1,"=",cell3.text.lower(),"   ")
					print(" ")
					print(" ")
					print("                                                                     ",u"\u2639")
					speaker.say(cell1)
					speaker.runAndWait()
					ws[str.format("E"+str(i))]=0
					wb.save(a)
					ws[str.format("C"+str(i))]=cell3.text.lower()
					wb.save(a)
					print(" ")
					print(" ")
					print("                                                                      Click enter to continue")
					clear=input()
					print("\033c", end="")
			
					
